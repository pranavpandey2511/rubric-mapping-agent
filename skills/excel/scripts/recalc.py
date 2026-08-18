#!/usr/bin/env python3
"""Create a recalculated XLSX copy without modifying the source workbook."""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


FORMULA_ERRORS = (
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
)


def _find_soffice() -> str | None:
    configured = os.getenv("RUBRIC_MAP_LIBREOFFICE_BIN")
    if configured:
        path = Path(configured).expanduser()
        return str(path) if path.is_file() else None
    for name in ("soffice", "libreoffice"):
        executable = shutil.which(name)
        if executable:
            return executable
    macos = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    return str(macos) if macos.is_file() else None


def _has_external_links(path: Path) -> bool:
    try:
        with ZipFile(path) as archive:
            return any(
                name.startswith("xl/externalLinks/")
                for name in archive.namelist()
            )
    except (BadZipFile, OSError):
        return False


def _formula_map(path: Path) -> dict[tuple[str, str], str]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        return {
            (worksheet.title, cell.coordinate): cell.value
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        }
    finally:
        workbook.close()


def _cache_summary(
    path: Path,
    formula_cells: dict[tuple[str, str], str],
) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        missing: list[str] = []
        errors: dict[str, list[str]] = {}
        for sheet, coordinate in formula_cells:
            value = workbook[sheet][coordinate].value
            location = f"{sheet}!{coordinate}"
            if value is None:
                missing.append(location)
            elif isinstance(value, str):
                for error in FORMULA_ERRORS:
                    if error in value:
                        errors.setdefault(error, []).append(location)
                        break
        return {
            "formula_count": len(formula_cells),
            "cached_formula_count": len(formula_cells) - len(missing),
            "missing_cache_count": len(missing),
            "missing_cache_cells": missing[:100],
            "formula_errors": {
                error: locations[:100]
                for error, locations in sorted(errors.items())
            },
        }
    finally:
        workbook.close()


def recalculate_copy(
    source: Path,
    output: Path,
    *,
    timeout_seconds: float = 60.0,
) -> dict[str, Any]:
    """Recalculate ``source`` into a distinct ``output`` XLSX path."""

    source = source.expanduser().resolve()
    output = output.expanduser().resolve()
    if not source.is_file():
        return {"error": f"Source workbook does not exist: {source}"}
    if source.suffix.lower() != ".xlsx" or output.suffix.lower() != ".xlsx":
        return {"error": "Recalculation supports .xlsx source and output files only"}
    if source == output:
        return {"error": "Output must differ from the source workbook"}
    if output.exists():
        return {"error": f"Refusing to overwrite existing output: {output}"}
    if _has_external_links(source):
        return {
            "error": (
                "Refusing to recalculate a workbook with external links; "
                "LibreOffice may replace unresolved linked values or formulas"
            )
        }

    soffice = _find_soffice()
    if soffice is None:
        return {"error": "LibreOffice soffice is not available in this environment"}

    source_formulas = _formula_map(source)
    with tempfile.TemporaryDirectory(prefix="xlsx-recalc-") as temporary:
        temporary_root = Path(temporary)
        converted_dir = temporary_root / "converted"
        profile_dir = temporary_root / "profile"
        converted_dir.mkdir()
        profile_dir.mkdir()
        command = [
            soffice,
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--headless",
            "--nologo",
            "--norestore",
            "--nodefault",
            "--nofirststartwizard",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            "--outdir",
            str(converted_dir),
            str(source),
        ]
        try:
            completed = subprocess.run(
                command,
                check=False,
                capture_output=True,
                text=True,
                timeout=timeout_seconds,
            )
        except subprocess.TimeoutExpired:
            return {
                "error": (
                    f"LibreOffice timed out after {timeout_seconds:g} seconds"
                )
            }
        if completed.returncode != 0:
            detail = completed.stderr.strip() or completed.stdout.strip()
            return {
                "error": (
                    f"LibreOffice exited with status {completed.returncode}: {detail}"
                )
            }

        converted = converted_dir / source.name
        if not converted.is_file():
            return {"error": "LibreOffice did not produce a recalculated workbook"}
        converted_formulas = _formula_map(converted)
        if converted_formulas != source_formulas:
            changed = sorted(
                set(source_formulas.items()) ^ set(converted_formulas.items())
            )
            return {
                "error": "LibreOffice changed formula text; recalculated copy rejected",
                "formula_differences": [
                    f"{sheet}!{coordinate}: {formula}"
                    for (sheet, coordinate), formula in changed[:100]
                ],
            }

        summary = _cache_summary(converted, converted_formulas)
        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(converted, output)

    return {
        "status": "success",
        "source": str(source),
        "output": str(output),
        **summary,
    }


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Create a recalculated XLSX copy using LibreOffice"
    )
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--timeout", type=float, default=60.0)
    return parser


def main() -> int:
    arguments = _parser().parse_args()
    result = recalculate_copy(
        arguments.source,
        arguments.output,
        timeout_seconds=arguments.timeout,
    )
    print(json.dumps(result, indent=2, sort_keys=True))
    return 1 if "error" in result else 0


if __name__ == "__main__":
    raise SystemExit(main())
