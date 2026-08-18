"""Validation and combination of stage artifacts."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from rubric_mapping_eval.i2c_mapping import parse_item_mapping
from rubric_mapping_eval.sectioning import parse_sections

from .handoff import (
    SummaryRecord,
    parse_summary_records,
    render_section_summary,
    validate_section_summary,
)
from .retrieval_index import (
    INDEX_GENERATOR,
    INDEX_SCHEMA_VERSION,
    validate_subsection_index,
)


def require_files(paths: Iterable[Path]) -> None:
    missing = [path for path in paths if not path.is_file()]
    if missing:
        raise FileNotFoundError("Missing input file(s): " + ", ".join(map(str, missing)))


def workbook_sheet_names(input_path: Path, complete_path: Path) -> tuple[str, ...]:
    """Return complete-workbook order after checking exact worksheet parity."""

    require_files((input_path, complete_path))
    input_workbook = load_workbook(input_path, read_only=True, data_only=False)
    complete_workbook = load_workbook(complete_path, read_only=True, data_only=False)
    try:
        input_names = tuple(input_workbook.sheetnames)
        complete_names = tuple(complete_workbook.sheetnames)
    finally:
        input_workbook.close()
        complete_workbook.close()

    if set(input_names) != set(complete_names):
        missing_from_input = sorted(set(complete_names) - set(input_names))
        missing_from_complete = sorted(set(input_names) - set(complete_names))
        raise ValueError(
            "Sheet-scoped execution requires matching worksheet names; "
            f"missing from input={missing_from_input}, "
            f"missing from complete={missing_from_complete}"
        )
    return complete_names


def combine_part1_artifacts(
    scoped_artifacts: Iterable[tuple[str | None, dict[str, Any]]],
    *,
    allowed_sheets: set[str],
) -> tuple[dict[str, Any], str]:
    """Combine scoped mappings and descriptions under final section IDs."""

    combined: list[dict[str, Any]] = []
    combined_summaries: list[SummaryRecord] = []
    for target_sheet, artifact in scoped_artifacts:
        scope_label = target_sheet if target_sheet is not None else "workbook"
        if set(artifact) != {"sections", "section_summaries"}:
            raise ValueError(
                f"Part 1 scope {scope_label!r} must return sections and "
                "section_summaries"
            )
        sections = parse_sections(
            {"sections": artifact["sections"]},
            context=f"Part 1 scope {scope_label!r}",
            allow_empty_sections=True,
        )
        if target_sheet is not None and any(
            section.sheet != target_sheet for section in sections
        ):
            raise ValueError(
                f"Part 1 worksheet {target_sheet!r} returned a section for another sheet"
            )
        invalid_sheets = sorted(
            {section.sheet for section in sections if section.sheet not in allowed_sheets}
        )
        if invalid_sheets:
            raise ValueError(
                f"Part 1 returned sections for unknown worksheets: {invalid_sheets}"
            )
        summaries = parse_summary_records(
            artifact["section_summaries"],
            id_field="section_id",
            expected_ids=[section.section_id for section in sections],
            context=f"Part 1 scope {scope_label!r} section_summaries",
        )
        for section, summary in zip(sections, summaries, strict=True):
            final_id = f"section_{len(combined) + 1:03d}"
            combined.append(
                {
                    "section_id": final_id,
                    "sheet": section.sheet,
                    "cells": [cell.address for cell in section.cells],
                }
            )
            combined_summaries.append(
                SummaryRecord(
                    identifier=final_id,
                    title=summary.title,
                    detail=summary.detail,
                    plain_language=summary.plain_language,
                )
            )
    payload = {"sections": combined}
    summary_text = render_section_summary(combined, combined_summaries)
    validate_section_summary(summary_text, combined)
    return payload, summary_text


def combine_part2_artifacts(
    scoped_artifacts: Iterable[tuple[str | None, dict[str, Any]]],
    sections_path: Path,
    eligible: set[tuple[str, str]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Concatenate agent-authored Part 2 artifacts without rewriting them."""

    combined: list[dict[str, Any]] = []
    combined_families: list[dict[str, Any]] = []
    combined_relationships: list[dict[str, Any]] = []
    for target_sheet, artifact in scoped_artifacts:
        scope_label = target_sheet if target_sheet is not None else "workbook"
        if set(artifact) != {"subsections", "subsection_index"}:
            raise ValueError(
                f"Part 2 scope {scope_label!r} must return subsections and "
                "subsection_index"
            )
        local_payload = {"subsections": artifact["subsections"]}
        validate_subsections(local_payload, sections_path)
        local_subsections = local_payload["subsections"]
        if target_sheet is not None and any(
            subsection["sheet"] != target_sheet for subsection in local_subsections
        ):
            raise ValueError(
                f"Part 2 worksheet {target_sheet!r} returned a subsection for another sheet"
            )
        local_index = artifact["subsection_index"]
        validate_subsection_index(
            local_index,
            subsections=local_subsections,
            eligible=eligible,
        )
        combined.extend(local_subsections)
        combined_families.extend(local_index["families"])
        combined_relationships.extend(local_index["relationships"])

    payload = {"subsections": combined}
    validate_subsections(payload, sections_path)
    index_payload = {
        "schema_version": INDEX_SCHEMA_VERSION,
        "generated_by": INDEX_GENERATOR,
        "families": combined_families,
        "relationships": combined_relationships,
    }
    validate_subsection_index(
        index_payload,
        subsections=combined,
        eligible=eligible,
    )
    return payload, index_payload


def build_part2_artifacts(
    artifact: dict[str, Any],
    sections_path: Path,
    eligible: set[tuple[str, str]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Validate the two direct agent-authored Part 2 artifacts."""

    if set(artifact) != {"subsections", "subsection_index"}:
        raise ValueError("Part 2 must return subsections and subsection_index")
    payload = {"subsections": artifact["subsections"]}
    validate_subsections(payload, sections_path)
    index_payload = artifact["subsection_index"]
    validate_subsection_index(
        index_payload,
        subsections=payload["subsections"],
        eligible=eligible,
    )
    return payload, index_payload


def validate_subsections(
    payload: dict[str, Any], sections_path: Path | None = None
) -> None:
    if (
        set(payload) != {"subsections"}
        or not isinstance(payload["subsections"], list)
        or not payload["subsections"]
    ):
        raise ValueError("subsections.json must contain only a subsections list")
    parents = None
    if sections_path is not None:
        sections = parse_sections(json.loads(sections_path.read_text(encoding="utf-8")))
        parents = {section.section_id: section for section in sections}
    seen: set[str] = set()
    required = {"subsection_id", "parent_section_id", "sheet", "cells", "roles"}
    for index, subsection in enumerate(payload["subsections"]):
        if not isinstance(subsection, dict) or set(subsection) != required:
            raise ValueError(f"subsections[{index}] must contain exactly {sorted(required)}")
        identifier = subsection["subsection_id"]
        if not isinstance(identifier, str) or not identifier or identifier in seen:
            raise ValueError(f"invalid or duplicate subsection_id at index {index}")
        seen.add(identifier)
        for field in ("parent_section_id", "sheet"):
            if not isinstance(subsection[field], str) or not subsection[field]:
                raise ValueError(f"subsections[{index}].{field} must be non-empty")
        if (
            not isinstance(subsection["cells"], list)
            or not subsection["cells"]
            or not all(isinstance(value, str) and value for value in subsection["cells"])
        ):
            raise ValueError(f"subsections[{index}].cells must contain addresses")
        if (
            not isinstance(subsection["roles"], list)
            or not subsection["roles"]
            or not all(isinstance(value, str) and value for value in subsection["roles"])
        ):
            raise ValueError(f"subsections[{index}].roles must contain semantic tags")
        if parents is not None:
            parent = parents.get(subsection["parent_section_id"])
            if parent is None or parent.sheet != subsection["sheet"]:
                raise ValueError(f"subsections[{index}] has an invalid Part 1 parent")
            parent_addresses = {cell.address for cell in parent.cells}
            if not set(subsection["cells"]).issubset(parent_addresses):
                raise ValueError(f"subsections[{index}] contains cells outside its parent")


def eligible_diff_cells(
    input_path: Path, complete_path: Path
) -> set[tuple[str, str]]:
    initial = load_workbook(input_path, data_only=False, read_only=False)
    complete = load_workbook(complete_path, data_only=False, read_only=False)
    try:
        eligible: set[tuple[str, str]] = set()
        for completed_sheet in complete.worksheets:
            initial_sheet = (
                initial[completed_sheet.title]
                if completed_sheet.title in initial.sheetnames
                else None
            )
            for row in completed_sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    initial_value = (
                        initial_sheet[cell.coordinate].value
                        if initial_sheet is not None
                        else None
                    )
                    if initial_value != cell.value:
                        eligible.add((completed_sheet.title, cell.coordinate.upper()))
        return eligible
    finally:
        initial.close()
        complete.close()


def validate_part3_artifact(
    artifact: dict[str, Any],
    *,
    expected_item_ids: tuple[str, ...],
    eligible: set[tuple[str, str]],
    target_sheet: str | None = None,
) -> dict[str, frozenset[Any]]:
    parsed = parse_item_mapping(artifact, allow_empty_cells=True)
    if set(parsed) != set(expected_item_ids):
        raise ValueError("items_to_cells.json item IDs do not match rubric.json")
    if target_sheet is not None:
        wrong_sheet = sorted(
            (cell.sheet, cell.address)
            for cells in parsed.values()
            for cell in cells
            if cell.sheet != target_sheet
        )
        if wrong_sheet:
            raise ValueError(
                f"Part 3 worksheet {target_sheet!r} returned cells for another sheet: "
                f"{wrong_sheet[:10]}"
            )
    invalid = sorted(
        (cell.sheet, cell.address)
        for cells in parsed.values()
        for cell in cells
        if (cell.sheet, cell.address) not in eligible
    )
    if invalid:
        raise ValueError(f"Part 3 mapped cells outside the eligible diff: {invalid[:10]}")
    return parsed


def combine_part3_artifacts(
    sheet_artifacts: Iterable[tuple[str, dict[str, Any]]],
    *,
    expected_item_ids: tuple[str, ...],
    eligible: set[tuple[str, str]],
    workbook_sheet_order: tuple[str, ...],
) -> dict[str, Any]:
    """Union sheet-scoped item mappings into the public workbook artifact."""

    combined = {item_id: set() for item_id in expected_item_ids}
    for target_sheet, artifact in sheet_artifacts:
        parsed = validate_part3_artifact(
            artifact,
            expected_item_ids=expected_item_ids,
            eligible=eligible,
            target_sheet=target_sheet,
        )
        for item_id, cells in parsed.items():
            combined[item_id].update(cells)
    sheet_position = {
        sheet_name: index for index, sheet_name in enumerate(workbook_sheet_order)
    }
    return {
        "items": [
            {
                "item_id": item_id,
                "cells": [
                    cell.to_dict()
                    for cell in sorted(
                        combined[item_id],
                        key=lambda cell: (
                            sheet_position[cell.sheet],
                            cell.address,
                        ),
                    )
                ],
            }
            for item_id in expected_item_ids
        ]
    }
