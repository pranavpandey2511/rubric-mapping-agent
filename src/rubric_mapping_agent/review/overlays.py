"""Apply Parts 1-3 mapping overlays to a review workbook."""

from __future__ import annotations

from collections import defaultdict
from typing import Any, Iterable

from openpyxl.comments import Comment
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
from rubric_mapping_eval.common import CellRef
from rubric_mapping_eval.sectioning import Section

from .contracts import RubricItemDetail, Subsection


COMMENT_MARKER = "--- Rubric Mapping Visualization ---"
COMMENT_AUTHOR = "Rubric Mapping Visualizer"

# Part 1 is deliberately uniform; Part 2 is the only multi-color boundary layer.
PART1_OUTLINE_COLOR = "1F4E78"
PART1_LEGEND_FILL = "D9EAF7"
PART2_HISTORICAL_COLOR = "2F75B5"
PART2_PROJECTED_COLOR = "ED7D31"
PART3_HIGHLIGHT_COLOR = "EDE9FE"
SECTION_PALETTE = (
    ("2563EB", "DBEAFE"),
    ("DC2626", "FEE2E2"),
    ("16A34A", "DCFCE7"),
    ("9333EA", "F3E8FF"),
    ("EA580C", "FFEDD5"),
    ("0891B2", "CFFAFE"),
    ("CA8A04", "FEF9C3"),
    ("DB2777", "FCE7F3"),
    ("4F46E5", "E0E7FF"),
    ("0F766E", "CCFBF1"),
    ("7C3AED", "EDE9FE"),
    ("475569", "E2E8F0"),
)
FALLBACK_COLOR = ("64748B", "E2E8F0")


def color_map(section_ids: Iterable[str]) -> dict[str, tuple[str, str]]:
    return {
        section_id: SECTION_PALETTE[index % len(SECTION_PALETTE)]
        for index, section_id in enumerate(dict.fromkeys(section_ids))
    }


def _replace_border_side(border: Border, name: str, side: Side) -> Border:
    values = {
        "left": border.left,
        "right": border.right,
        "top": border.top,
        "bottom": border.bottom,
        "diagonal": border.diagonal,
        "diagonal_direction": border.diagonal_direction,
        "vertical": border.vertical,
        "horizontal": border.horizontal,
        "diagonalUp": border.diagonalUp,
        "diagonalDown": border.diagonalDown,
        "outline": border.outline,
        "start": border.start,
        "end": border.end,
    }
    values[name] = side
    return Border(**values)


def apply_subsection_underlines(
    workbook,
    subsections: tuple[Subsection, ...],
    colors: dict[str, tuple[str, str]],
    sections_by_id: dict[str, Section],
) -> None:
    period_headers_by_section: dict[
        str, tuple[tuple[int, tuple[tuple[int, int, str], ...]], ...]
    ] = {}

    def period_kind(value: Any) -> str | None:
        if not isinstance(value, str):
            return None
        normalized = value.strip().lower()
        if normalized in {"historical", "actual", "actuals"}:
            return "historical"
        if normalized in {
            "projected",
            "projection",
            "forecast",
            "forecasted",
            "budget",
            "outlook",
        }:
            return "projected"
        return None

    def merged_column_span(
        sheet: Worksheet, row: int, column: int
    ) -> tuple[int, int]:
        address = sheet.cell(row=row, column=column).coordinate
        for merged_range in sheet.merged_cells.ranges:
            if address in merged_range:
                return merged_range.min_col, merged_range.max_col
        return column, column

    def period_headers(
        section: Section,
    ) -> tuple[tuple[int, tuple[tuple[int, int, str], ...]], ...]:
        cached = period_headers_by_section.get(section.section_id)
        if cached is not None:
            return cached
        sheet = workbook[section.sheet]
        coordinates = {coordinate_to_tuple(cell.address) for cell in section.cells}
        if not coordinates:
            return ()
        rows = sorted({row for row, _ in coordinates})
        columns = sorted({column for _, column in coordinates})
        detected: list[tuple[int, tuple[tuple[int, int, str], ...]]] = []
        for row in rows:
            bands: list[tuple[int, int, str]] = []
            for column in columns:
                if (row, column) not in coordinates:
                    continue
                kind = period_kind(sheet.cell(row=row, column=column).value)
                if kind is None:
                    continue
                start_column, end_column = merged_column_span(sheet, row, column)
                bands.append((start_column, end_column, kind))
            if bands:
                detected.append((row, tuple(sorted(set(bands)))))
        result = tuple(detected)
        period_headers_by_section[section.section_id] = result
        return result

    def row_period_bands(
        headers: tuple[tuple[int, tuple[tuple[int, int, str], ...]], ...],
        row: int,
    ) -> tuple[tuple[int, int, str], ...] | None:
        if any(header_row == row for header_row, _ in headers):
            return None
        applicable = [bands for header_row, bands in headers if header_row < row]
        return applicable[-1] if applicable else None

    for subsection in subsections:
        sheet = workbook[subsection.sheet]
        parent = sections_by_id[subsection.parent_section_id]
        headers = period_headers(parent)
        fallback, _ = colors.get(subsection.subsection_id, FALLBACK_COLOR)
        normalized_roles = " ".join(subsection.roles).lower()
        if "historical" in normalized_roles and "project" not in normalized_roles:
            fallback = PART2_HISTORICAL_COLOR
        elif any(
            term in normalized_roles
            for term in ("projected", "projection", "forecast", "budget", "outlook")
        ) and "historical" not in normalized_roles:
            fallback = PART2_PROJECTED_COLOR
        coordinates = {coordinate_to_tuple(address) for address in subsection.cells}
        for row in sorted({coordinate[0] for coordinate in coordinates}):
            bands = row_period_bands(headers, row)
            if headers and bands is None:
                # Metadata above the first time-series header is not a row boundary.
                continue
            row_coordinates = sorted(
                coordinate for coordinate in coordinates if coordinate[0] == row
            )
            for _, column in row_coordinates:
                cell = sheet.cell(row=row, column=column)
                if cell.value in (None, ""):
                    continue
                color = fallback
                if bands is not None:
                    matching_band = next(
                        (
                            kind
                            for start_column, end_column, kind in bands
                            if start_column <= column <= end_column
                        ),
                        None,
                    )
                    if matching_band is None:
                        continue
                    color = (
                        PART2_HISTORICAL_COLOR
                        if matching_band == "historical"
                        else PART2_PROJECTED_COLOR
                    )
                underline = Side(style="medium", color=color)
                cell.border = _replace_border_side(cell.border, "bottom", underline)


def apply_section_outlines(workbook, sections: tuple[Section, ...]) -> None:
    for section in sections:
        sheet = workbook[section.sheet]
        coordinates = {coordinate_to_tuple(cell.address) for cell in section.cells}
        outline = Side(style="medium", color=PART1_OUTLINE_COLOR)
        for row, column in coordinates:
            cell = sheet.cell(row=row, column=column)
            if (row - 1, column) not in coordinates:
                cell.border = _replace_border_side(cell.border, "top", outline)
            if (row + 1, column) not in coordinates:
                cell.border = _replace_border_side(cell.border, "bottom", outline)
            if (row, column - 1) not in coordinates:
                cell.border = _replace_border_side(cell.border, "left", outline)
            if (row, column + 1) not in coordinates:
                cell.border = _replace_border_side(cell.border, "right", outline)


def _merged_anchor(sheet: Worksheet, address: str):
    for merged_range in sheet.merged_cells.ranges:
        if address in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col)
    return sheet[address]


def _comment_text(
    item_ids: list[str],
    section_ids: tuple[str, ...],
    details: dict[str, RubricItemDetail],
) -> str:
    lines = [COMMENT_MARKER, "Part 3 rubric item mapping"]
    if section_ids:
        lines.append("Part 1 section(s): " + ", ".join(section_ids))
    for item_id in item_ids:
        detail = details.get(item_id)
        if detail is None:
            lines.append(f"- {item_id}")
            continue
        criterion = f"criterion {detail.criterion}"
        if detail.criterion_description:
            criterion += f" — {detail.criterion_description}"
        lines.append(f"- {item_id} ({criterion})")
        if detail.condition:
            lines.append(f"  {detail.condition}")
    return "\n".join(lines)[:32_000]


def _set_generated_comment(cell, generated_text: str) -> None:
    original = cell.comment
    if original is None:
        cell.comment = Comment(generated_text, COMMENT_AUTHOR)
        return
    preserved = original.text.split(COMMENT_MARKER, 1)[0].rstrip()
    combined = f"{preserved}\n\n{generated_text}" if preserved else generated_text
    cell.comment = Comment(combined[:32_000], original.author or COMMENT_AUTHOR)


def apply_item_highlights(
    workbook,
    items: dict[str, frozenset[CellRef]],
    sections_by_cell: dict[tuple[str, str], tuple[str, ...]],
    details: dict[str, RubricItemDetail],
) -> None:
    items_by_cell: dict[tuple[str, str], list[str]] = defaultdict(list)
    for item_id, cells in items.items():
        for cell in sorted(cells):
            items_by_cell[(cell.sheet, cell.address)].append(item_id)

    comments_by_anchor: dict[tuple[str, str], list[str]] = defaultdict(list)
    sections_by_anchor: dict[tuple[str, str], list[str]] = defaultdict(list)
    for (sheet_name, address), item_ids in items_by_cell.items():
        sheet = workbook[sheet_name]
        section_ids = sections_by_cell.get((sheet_name, address), ())
        highlight = PatternFill(
            fill_type="solid",
            fgColor=PART3_HIGHLIGHT_COLOR,
        )
        merged_range = next(
            (
                cell_range
                for cell_range in sheet.merged_cells.ranges
                if address in cell_range
            ),
            None,
        )
        if merged_range is None:
            sheet[address].fill = highlight
        else:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for column in range(merged_range.min_col, merged_range.max_col + 1):
                    sheet.cell(row=row, column=column).fill = highlight
        comment_cell = _merged_anchor(sheet, address)
        anchor = (sheet_name, comment_cell.coordinate)
        comments_by_anchor[anchor].extend(
            item_id for item_id in item_ids if item_id not in comments_by_anchor[anchor]
        )
        sections_by_anchor[anchor].extend(
            section_id
            for section_id in section_ids
            if section_id not in sections_by_anchor[anchor]
        )

    for (sheet_name, address), item_ids in comments_by_anchor.items():
        comment_cell = workbook[sheet_name][address]
        _set_generated_comment(
            comment_cell,
            _comment_text(
                item_ids,
                tuple(sections_by_anchor[(sheet_name, address)]),
                details,
            ),
        )
