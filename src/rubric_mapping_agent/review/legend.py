"""Build the explanatory worksheet for annotated review workbooks."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rubric_mapping_eval.common import CellRef
from rubric_mapping_eval.sectioning import Section

from .contracts import RubricItemDetail, Subsection
from .overlays import (
    FALLBACK_COLOR,
    PART1_LEGEND_FILL,
    PART1_OUTLINE_COLOR,
    PART2_HISTORICAL_COLOR,
    PART2_PROJECTED_COLOR,
    PART3_HIGHLIGHT_COLOR,
)


LEGEND_TITLE = "Mapping Legend"
LEGEND_MARKER = "Rubric Mapping Visualization"


def _legend_sheet(workbook):
    if LEGEND_TITLE in workbook.sheetnames:
        existing = workbook[LEGEND_TITLE]
        if existing["A1"].value == LEGEND_MARKER:
            workbook.remove(existing)
        else:
            index = 2
            while f"{LEGEND_TITLE} {index}" in workbook.sheetnames:
                index += 1
            return workbook.create_sheet(f"{LEGEND_TITLE} {index}")
    return workbook.create_sheet(LEGEND_TITLE)


def add_legend(
    workbook,
    sections: tuple[Section, ...],
    subsections: tuple[Subsection, ...],
    items: dict[str, frozenset[CellRef]],
    subsection_colors: dict[str, tuple[str, str]],
    details: dict[str, RubricItemDetail],
) -> None:
    sheet = _legend_sheet(workbook)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.merge_cells("A1:E1")
    sheet["A1"] = LEGEND_MARKER
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="0F172A")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells("A2:E2")
    sheet["A2"] = (
        "Part 1 = uniform navy outline; Part 2 = row-level blue historical and "
        "orange projected boundaries (other blocks use subsection colors); Part 3 "
        "= full-cell lavender highlight with an Excel comment."
    )
    sheet["A2"].font = Font(color="475569", italic=True)

    headers = ("Layer", "Color sample", "Identifier", "Sheet / parent", "Meaning")
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="334155")

    row = 5
    for section in sections:
        values = (
            "Part 1",
            "uniform outline",
            section.section_id,
            section.sheet,
            "overall section",
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
        sheet.cell(row=row, column=2).fill = PatternFill(
            fill_type="solid", fgColor=PART1_LEGEND_FILL
        )
        sheet.cell(row=row, column=2).border = Border(
            left=Side(style="medium", color=PART1_OUTLINE_COLOR),
            right=Side(style="medium", color=PART1_OUTLINE_COLOR),
            top=Side(style="medium", color=PART1_OUTLINE_COLOR),
            bottom=Side(style="medium", color=PART1_OUTLINE_COLOR),
        )
        row += 1

    for subsection in subsections:
        dark, _ = subsection_colors.get(subsection.subsection_id, FALLBACK_COLOR)
        values = (
            "Part 2",
            "row-level boundaries",
            subsection.subsection_id,
            subsection.parent_section_id,
            ", ".join(subsection.roles),
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
        sample = sheet.cell(row=row, column=2)
        if "review-only-fallback" in subsection.roles:
            sample.border = Border(bottom=Side(style="medium", color=dark))
        else:
            sample.border = Border(
                top=Side(style="medium", color=PART2_HISTORICAL_COLOR),
                bottom=Side(style="medium", color=PART2_PROJECTED_COLOR),
            )
        row += 1

    for item_id, mapped_cells in items.items():
        detail = details.get(item_id)
        meaning = (
            detail.criterion_description
            if detail and detail.criterion_description
            else "mapped rubric item"
        )
        values = (
            "Part 3",
            "highlight + comment",
            item_id,
            f"{len(mapped_cells)} mapped cell(s)",
            meaning,
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
        sheet.cell(row=row, column=2).fill = PatternFill(
            fill_type="solid", fgColor=PART3_HIGHLIGHT_COLOR
        )
        row += 1

    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 34
    sheet.column_dimensions["E"].width = 80
    for cells in sheet.iter_rows(
        min_row=4, max_row=max(row - 1, 4), min_col=1, max_col=5
    ):
        for cell in cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
