"""Read-only workbook screenshot tool and viewport navigation runtime."""

from __future__ import annotations

import base64
import hashlib
import json
import os
import platform
import shutil
import socket
import struct
import subprocess
import tempfile
import threading
import time
from contextlib import nullcontext
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal, Protocol
from uuid import uuid4

from langchain_core.tools import StructuredTool, ToolException
from openpyxl.utils.cell import (
    coordinate_to_tuple,
    get_column_letter,
    range_boundaries,
)

from ..artifacts import write_json


VISUAL_BACKENDS = {"off", "libreoffice_pdf", "libreoffice_ui"}
NAVIGATION_ACTIONS = {
    "absolute",
    "page_down",
    "page_up",
    "page_left",
    "page_right",
}
BRIDGE_PATH = Path(__file__).with_name("_libreoffice_bridge.py")
_LIBREOFFICE_UI_CAPTURE_LOCK = threading.Lock()


class VisualBackendUnavailable(RuntimeError):
    """Raised when a requested local rendering dependency is unavailable."""


@dataclass(frozen=True)
class VisualWorkbookConfig:
    """Process and viewport configuration loaded at the orchestration boundary."""

    backend: Literal["off", "libreoffice_pdf", "libreoffice_ui"] = "off"
    width: int = 1440
    height: int = 900
    timeout_seconds: float = 45.0
    capture_delay_seconds: float = 0.6
    libreoffice_binary: str | None = None
    libreoffice_python: str | None = None
    pdf_rasterizer: str | None = None
    libreoffice_process_name: str = "LibreOffice"

    @classmethod
    def from_env(cls) -> "VisualWorkbookConfig":
        backend = os.getenv("RUBRIC_MAP_VISUAL_BACKEND", "off").strip().lower()
        if backend not in VISUAL_BACKENDS:
            choices = ", ".join(sorted(VISUAL_BACKENDS))
            raise ValueError(
                f"RUBRIC_MAP_VISUAL_BACKEND must be one of: {choices}"
            )
        return cls(
            backend=backend,  # type: ignore[arg-type]
            width=_bounded_env_int("RUBRIC_MAP_VISUAL_WIDTH", 1440, 640, 3840),
            height=_bounded_env_int("RUBRIC_MAP_VISUAL_HEIGHT", 900, 480, 2160),
            timeout_seconds=_bounded_env_float(
                "RUBRIC_MAP_VISUAL_TIMEOUT_SECONDS", 45.0, 5.0, 180.0
            ),
            capture_delay_seconds=_bounded_env_float(
                "RUBRIC_MAP_VISUAL_CAPTURE_DELAY_SECONDS", 0.6, 0.0, 5.0
            ),
            libreoffice_binary=os.getenv("RUBRIC_MAP_LIBREOFFICE_BIN") or None,
            libreoffice_python=os.getenv("RUBRIC_MAP_LIBREOFFICE_PYTHON") or None,
            pdf_rasterizer=os.getenv("RUBRIC_MAP_PDF_RASTERIZER") or None,
            libreoffice_process_name=os.getenv(
                "RUBRIC_MAP_LIBREOFFICE_PROCESS_NAME", "LibreOffice"
            ),
        )


def _bounded_env_int(name: str, default: int, minimum: int, maximum: int) -> int:
    raw = os.getenv(name)
    value = default if raw is None else int(raw)
    if not minimum <= value <= maximum:
        raise ValueError(f"{name} must be between {minimum} and {maximum}")
    return value


def _bounded_env_float(
    name: str,
    default: float,
    minimum: float,
    maximum: float,
) -> float:
    raw = os.getenv(name)
    value = default if raw is None else float(raw)
    if not minimum <= value <= maximum:
        raise ValueError(f"{name} must be between {minimum} and {maximum}")
    return value


@dataclass(frozen=True)
class ViewportRequest:
    workbook: str
    path: Path
    sheet: str
    cell_range: str
    zoom: int


@dataclass(frozen=True)
class ViewportCapture:
    png: bytes
    engine: str
    sheet: str
    requested_range: str
    visible_range: str
    zoom: int | None
    width: int
    height: int


class VisualBackend(Protocol):
    def capture(self, request: ViewportRequest) -> ViewportCapture: ...

    def close(self) -> None: ...


def _resolve_executable(
    configured: str | None,
    *,
    names: tuple[str, ...],
    candidates: tuple[Path, ...] = (),
) -> str:
    if configured:
        path = Path(configured).expanduser()
        if not path.is_file():
            raise VisualBackendUnavailable(f"Configured executable not found: {path}")
        return str(path)
    for name in names:
        found = shutil.which(name)
        if found:
            return found
    for candidate in candidates:
        if candidate.is_file():
            return str(candidate)
    rendered = ", ".join(names) or "configured executable"
    raise VisualBackendUnavailable(f"Could not locate {rendered}")


def _sanitized_process_env(*, headless: bool) -> dict[str, str]:
    sensitive_markers = ("KEY", "TOKEN", "SECRET", "PASSWORD", "CREDENTIAL")
    env = {
        key: value
        for key, value in os.environ.items()
        if not any(marker in key.upper() for marker in sensitive_markers)
    }
    if headless:
        env["SAL_USE_VCLPLUGIN"] = "svp"
    else:
        env.pop("SAL_USE_VCLPLUGIN", None)
    return env


class _LibreOfficeProcess:
    """One isolated LibreOffice process controlled through a local UNO socket."""

    def __init__(self, config: VisualWorkbookConfig, *, headless: bool) -> None:
        self.config = config
        self.headless = headless
        self.soffice = _resolve_executable(
            config.libreoffice_binary,
            names=("soffice", "libreoffice"),
            candidates=(
                Path("/Applications/LibreOffice.app/Contents/MacOS/soffice"),
            ),
        )
        self.python = _resolve_executable(
            config.libreoffice_python,
            names=(),
            candidates=(
                Path("/Applications/LibreOffice.app/Contents/Resources/python"),
                Path("/usr/lib/libreoffice/program/python"),
            ),
        )
        self._profile = tempfile.TemporaryDirectory(prefix="rubric-map-lo-profile-")
        self.port = _available_port()
        self.process: subprocess.Popen[bytes] | None = None

    def start(self) -> None:
        if self.process is not None:
            return
        profile_uri = Path(self._profile.name).resolve().as_uri()
        command = [
            self.soffice,
            f"-env:UserInstallation={profile_uri}",
            "--nologo",
            "--norestore",
            "--nodefault",
            "--nofirststartwizard",
            f"--accept=socket,host=127.0.0.1,port={self.port};urp;"
            "StarOffice.ComponentContext",
        ]
        if self.headless:
            command.append("--headless")
        self.process = subprocess.Popen(
            command,
            env=_sanitized_process_env(headless=self.headless),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        deadline = time.monotonic() + self.config.timeout_seconds
        last_error = "LibreOffice did not accept UNO connections"
        while time.monotonic() < deadline:
            if self.process.poll() is not None:
                raise VisualBackendUnavailable(
                    f"LibreOffice exited with code {self.process.returncode}"
                )
            try:
                self.run_bridge("ping", timeout_seconds=3.0)
                return
            except Exception as exc:
                last_error = str(exc)
                time.sleep(0.2)
        self.close()
        raise VisualBackendUnavailable(
            f"LibreOffice was not ready within {self.config.timeout_seconds}s: "
            f"{last_error}"
        )

    def run_bridge(
        self,
        command: str,
        *arguments: str,
        timeout_seconds: float | None = None,
    ) -> dict[str, Any]:
        if self.process is None:
            raise RuntimeError("LibreOffice has not been started")
        completed = subprocess.run(
            [
                self.python,
                str(BRIDGE_PATH),
                "--port",
                str(self.port),
                command,
                *arguments,
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=timeout_seconds or self.config.timeout_seconds,
            env=_sanitized_process_env(headless=self.headless),
        )
        if completed.returncode != 0:
            detail = completed.stderr.strip() or completed.stdout.strip()
            raise RuntimeError(f"LibreOffice {command} failed: {detail}")
        try:
            result = json.loads(completed.stdout)
        except json.JSONDecodeError as exc:
            raise RuntimeError(
                f"LibreOffice {command} returned invalid JSON"
            ) from exc
        if not isinstance(result, dict):
            raise RuntimeError(f"LibreOffice {command} returned a non-object result")
        return result

    def close(self) -> None:
        process = self.process
        self.process = None
        if process is not None and process.poll() is None:
            try:
                subprocess.run(
                    [
                        self.python,
                        str(BRIDGE_PATH),
                        "--port",
                        str(self.port),
                        "shutdown",
                    ],
                    check=False,
                    capture_output=True,
                    timeout=5,
                    env=_sanitized_process_env(headless=self.headless),
                )
            except Exception:
                pass
            try:
                process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                process.terminate()
                try:
                    process.wait(timeout=5)
                except subprocess.TimeoutExpired:
                    process.kill()
                    process.wait(timeout=5)
        self._profile.cleanup()


def _available_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as listener:
        listener.bind(("127.0.0.1", 0))
        return int(listener.getsockname()[1])


class LibreOfficePdfBackend:
    """Render an exact cell range through Calc PDF export and rasterization."""

    def __init__(self, config: VisualWorkbookConfig) -> None:
        self.config = config
        self._office: _LibreOfficeProcess | None = None
        self.rasterizer = _resolve_executable(
            config.pdf_rasterizer,
            names=("pdftocairo", "pdftoppm"),
        )
        self._temporary = tempfile.TemporaryDirectory(prefix="rubric-map-lo-render-")

    def _session(self) -> _LibreOfficeProcess:
        if self._office is None:
            office = _LibreOfficeProcess(self.config, headless=True)
            try:
                office.start()
            except Exception:
                office.close()
                raise
            self._office = office
        return self._office

    def capture(self, request: ViewportRequest) -> ViewportCapture:
        token = uuid4().hex
        output_dir = Path(self._temporary.name)
        pdf_path = output_dir / f"{token}.pdf"
        png_prefix = output_dir / token
        session = self._session()
        result = session.run_bridge(
            "render",
            "--workbook",
            str(request.path),
            "--sheet",
            request.sheet,
            "--cell-range",
            request.cell_range,
            "--output",
            str(pdf_path),
        )
        if not pdf_path.is_file() or pdf_path.stat().st_size == 0:
            raise RuntimeError("LibreOffice did not produce a PDF viewport")

        raster_name = Path(self.rasterizer).name
        raster_command = [
            self.rasterizer,
            "-png",
            "-singlefile",
            "-scale-to-x",
            str(self.config.width),
            "-scale-to-y",
            "-1",
            str(pdf_path),
            str(png_prefix),
        ]
        completed = subprocess.run(
            raster_command,
            check=False,
            capture_output=True,
            text=True,
            timeout=self.config.timeout_seconds,
        )
        if completed.returncode != 0:
            raise RuntimeError(
                f"{raster_name} failed: {completed.stderr.strip()}"
            )
        png_path = png_prefix.with_suffix(".png")
        if not png_path.is_file() or png_path.stat().st_size == 0:
            raise RuntimeError(f"{raster_name} did not produce a PNG viewport")
        png = png_path.read_bytes()
        width, height = _png_size(png)
        return ViewportCapture(
            png=png,
            engine="libreoffice_pdf",
            sheet=request.sheet,
            requested_range=f"{request.sheet}!{request.cell_range}",
            visible_range=str(
                result.get(
                    "visible_range", f"{request.sheet}!{request.cell_range}"
                )
            ),
            # PDF export fits the requested range to one page, so it has no
            # Calc UI zoom percentage.
            zoom=None,
            width=width,
            height=height,
        )

    def close(self) -> None:
        if self._office is not None:
            self._office.close()
            self._office = None
        self._temporary.cleanup()


class LibreOfficeUiBackend:
    """Position a real Calc window through UNO and capture that window on macOS."""

    def __init__(self, config: VisualWorkbookConfig) -> None:
        if platform.system() != "Darwin":
            raise VisualBackendUnavailable(
                "libreoffice_ui currently requires macOS; use "
                "libreoffice_pdf on headless hosts"
            )
        self.config = config
        self._office: _LibreOfficeProcess | None = None
        self.osascript = _resolve_executable(None, names=("osascript",))
        self.screencapture = _resolve_executable(None, names=("screencapture",))
        self._temporary = tempfile.TemporaryDirectory(prefix="rubric-map-lo-ui-")

    def _session(self) -> _LibreOfficeProcess:
        if self._office is None:
            office = _LibreOfficeProcess(self.config, headless=False)
            try:
                office.start()
            except Exception:
                office.close()
                raise
            self._office = office
        return self._office

    def _window_bounds(self) -> tuple[int, int, int, int]:
        process_name = self.config.libreoffice_process_name.replace('"', "")
        script = f'''
tell application "LibreOffice" to activate
delay 0.1
tell application "System Events"
  tell process "{process_name}"
    set frontmost to true
    set position of front window to {{0, 25}}
    set size of front window to {{{self.config.width}, {self.config.height}}}
    set windowPosition to position of front window
    set windowSize to size of front window
  end tell
end tell
return (item 1 of windowPosition as text) & "," & ¬
       (item 2 of windowPosition as text) & "," & ¬
       (item 1 of windowSize as text) & "," & ¬
       (item 2 of windowSize as text)
'''
        completed = subprocess.run(
            [self.osascript, "-e", script],
            check=False,
            capture_output=True,
            text=True,
            timeout=self.config.timeout_seconds,
        )
        if completed.returncode != 0:
            raise VisualBackendUnavailable(
                "Could not access the LibreOffice window. Grant Accessibility "
                f"permission to the host process: {completed.stderr.strip()}"
            )
        try:
            x, y, width, height = (
                int(part.strip()) for part in completed.stdout.strip().split(",")
            )
        except Exception as exc:
            raise RuntimeError("Could not parse LibreOffice window bounds") from exc
        return x, y, width, height

    def capture(self, request: ViewportRequest) -> ViewportCapture:
        session = self._session()
        bridge_args = (
            "--workbook",
            str(request.path),
            "--sheet",
            request.sheet,
            "--cell-range",
            request.cell_range,
            "--zoom",
            str(request.zoom),
        )
        # The first call opens the document and creates its window. Resize the
        # window before querying the final visible range.
        session.run_bridge("position", *bridge_args)
        x, y, width, height = self._window_bounds()
        result = session.run_bridge("position", *bridge_args)
        if self.config.capture_delay_seconds:
            time.sleep(self.config.capture_delay_seconds)

        png_path = Path(self._temporary.name) / f"{uuid4().hex}.png"
        completed = subprocess.run(
            [
                self.screencapture,
                "-x",
                "-t",
                "png",
                f"-R{x},{y},{width},{height}",
                str(png_path),
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=self.config.timeout_seconds,
        )
        if completed.returncode != 0 or not png_path.is_file():
            raise VisualBackendUnavailable(
                "Could not capture the LibreOffice window. Grant Screen "
                f"Recording permission to the host process: {completed.stderr.strip()}"
            )
        png = png_path.read_bytes()
        image_width, image_height = _png_size(png)
        return ViewportCapture(
            png=png,
            engine="libreoffice_ui",
            sheet=request.sheet,
            requested_range=str(
                result.get(
                    "requested_range", f"{request.sheet}!{request.cell_range}"
                )
            ),
            visible_range=str(
                result.get(
                    "visible_range", f"{request.sheet}!{request.cell_range}"
                )
            ),
            zoom=int(result.get("zoom", request.zoom)),
            width=image_width,
            height=image_height,
        )

    def close(self) -> None:
        if self._office is not None:
            self._office.close()
            self._office = None
        self._temporary.cleanup()


def _png_size(png: bytes) -> tuple[int, int]:
    if len(png) < 24 or png[:8] != b"\x89PNG\r\n\x1a\n":
        raise ValueError("Visual backend returned invalid PNG data")
    return struct.unpack(">II", png[16:24])


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalized_range(
    *,
    anchor: str | None,
    cell_range: str | None,
    navigation: str,
    rows: int,
    columns: int,
    previous_range: str | None,
) -> str:
    if navigation not in NAVIGATION_ACTIONS:
        choices = ", ".join(sorted(NAVIGATION_ACTIONS))
        raise ToolException(f"navigation must be one of: {choices}")
    if not 1 <= rows <= 200:
        raise ToolException("rows must be between 1 and 200")
    if not 1 <= columns <= 50:
        raise ToolException("columns must be between 1 and 50")

    if navigation == "absolute":
        if cell_range:
            min_col, min_row, max_col, max_row = _range_bounds(cell_range)
        else:
            try:
                min_row, min_col = coordinate_to_tuple(anchor or "A1")
            except ValueError as exc:
                raise ToolException(f"Invalid anchor: {anchor!r}") from exc
            max_row = min_row + rows - 1
            max_col = min_col + columns - 1
    else:
        if anchor or cell_range:
            raise ToolException(
                "anchor and cell_range must be omitted for relative navigation"
            )
        if previous_range is None:
            raise ToolException(
                "Relative navigation requires an earlier view of this workbook sheet"
            )
        min_col, min_row, max_col, max_row = _range_bounds(previous_range)
        row_count = max_row - min_row + 1
        column_count = max_col - min_col + 1
        if navigation == "page_down":
            min_row = max(1, max_row - 2)
            max_row = min_row + row_count - 1
        elif navigation == "page_up":
            min_row = max(1, min_row - row_count + 3)
            max_row = min_row + row_count - 1
        elif navigation == "page_right":
            min_col = max(1, max_col)
            max_col = min_col + column_count - 1
        else:
            min_col = max(1, min_col - column_count + 1)
            max_col = min_col + column_count - 1

    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def _range_bounds(cell_range: str) -> tuple[int, int, int, int]:
    if "!" in cell_range:
        cell_range = cell_range.rsplit("!", 1)[1]
    try:
        boundaries = range_boundaries(cell_range)
    except ValueError as exc:
        raise ToolException(f"Invalid cell range: {cell_range!r}") from exc
    if not all(isinstance(value, int) and value >= 1 for value in boundaries):
        raise ToolException(
            f"Cell range must be a bounded A1 rectangle: {cell_range!r}"
        )
    return boundaries  # type: ignore[return-value]


class VisualWorkbookRuntime:
    """Stage-scoped workbook allowlist, navigation state, backend, and tool."""

    def __init__(
        self,
        config: VisualWorkbookConfig,
        workbooks: dict[str, Path],
        *,
        allowed_sheets: set[str] | None = None,
        artifacts_dir: Path | None = None,
        backend: VisualBackend | None = None,
    ) -> None:
        self.config = config
        self.workbooks = {
            name: path.resolve()
            for name, path in workbooks.items()
            if path.suffix.lower() in {".xlsx", ".xlsm"}
        }
        self.allowed_sheets = allowed_sheets
        self.artifacts_dir = (
            artifacts_dir.resolve() if artifacts_dir is not None else None
        )
        self._previous_ranges: dict[tuple[str, str], str] = {}
        self._capture_lock = threading.Lock()
        self._backend = backend
        if self.config.backend != "off" and not self.workbooks:
            raise ValueError("Visual inspection requires at least one supplied workbook")

    @property
    def enabled(self) -> bool:
        return self.config.backend != "off"

    def _get_backend(self) -> VisualBackend:
        if self._backend is None:
            if self.config.backend == "libreoffice_pdf":
                self._backend = LibreOfficePdfBackend(self.config)
            elif self.config.backend == "libreoffice_ui":
                self._backend = LibreOfficeUiBackend(self.config)
            else:
                raise RuntimeError("Visual inspection is disabled")
        return self._backend

    def tool(self) -> StructuredTool:
        if not self.enabled:
            raise RuntimeError("Visual inspection is disabled")

        def inspect_workbook_view(
            workbook: str,
            sheet: str,
            anchor: str | None = None,
            cell_range: str | None = None,
            navigation: str = "absolute",
            rows: int = 40,
            columns: int = 14,
            zoom: int = 100,
        ) -> list[dict[str, Any]]:
            """Visually inspect one read-only supplied workbook viewport.

            Use workbook ``input`` or ``complete``. For an absolute view, pass
            either an anchor plus rows/columns or an exact cell_range. For a
            follow-up view, omit anchor/cell_range and use page_down, page_up,
            page_left, or page_right. The returned image is supporting visual
            evidence; use Python workbook structure for exact cell membership.
            """

            try:
                path = self.workbooks[workbook]
            except KeyError as exc:
                available = ", ".join(sorted(self.workbooks))
                raise ToolException(
                    f"Unknown workbook {workbook!r}; available: {available}"
                ) from exc
            if self.allowed_sheets is not None and sheet not in self.allowed_sheets:
                allowed = ", ".join(sorted(self.allowed_sheets))
                raise ToolException(
                    f"Worksheet {sheet!r} is outside this stage scope; allowed: {allowed}"
                )
            if not 20 <= zoom <= 400:
                raise ToolException("zoom must be between 20 and 400")

            key = (workbook, sheet)
            planned_range = _normalized_range(
                anchor=anchor,
                cell_range=cell_range,
                navigation=navigation,
                rows=rows,
                columns=columns,
                previous_range=self._previous_ranges.get(key),
            )
            request = ViewportRequest(
                workbook=workbook,
                path=path,
                sheet=sheet,
                cell_range=planned_range,
                zoom=zoom,
            )
            try:
                capture_context = (
                    _LIBREOFFICE_UI_CAPTURE_LOCK
                    if self.config.backend == "libreoffice_ui"
                    else (
                        self._capture_lock
                        if self.config.backend == "libreoffice_pdf"
                        else nullcontext()
                    )
                )
                with capture_context:
                    capture = self._get_backend().capture(request)
            except ToolException:
                raise
            except Exception as exc:
                raise ToolException(
                    f"Visual workbook inspection failed: {exc}"
                ) from exc

            visible_without_sheet = capture.visible_range.rsplit("!", 1)[-1]
            self._previous_ranges[key] = visible_without_sheet
            source_hash = _sha256(path)
            screenshot_hash = hashlib.sha256(capture.png).hexdigest()
            metadata = {
                "workbook": workbook,
                "sheet": capture.sheet,
                "requested_range": capture.requested_range,
                "visible_range": capture.visible_range,
                "zoom": capture.zoom,
                "zoom_mode": (
                    "calc_percentage" if capture.zoom is not None else "fit_range"
                ),
                "engine": capture.engine,
                "image_width": capture.width,
                "image_height": capture.height,
                "source_sha256": source_hash,
                "screenshot_sha256": screenshot_hash,
            }
            if self.artifacts_dir is not None:
                capture_id = uuid4().hex
                self.artifacts_dir.mkdir(parents=True, exist_ok=True)
                png_path = self.artifacts_dir / f"{capture_id}.png"
                metadata_path = self.artifacts_dir / f"{capture_id}.json"
                metadata.update(
                    {
                        "capture_id": capture_id,
                        "artifact_png": png_path.name,
                        "artifact_metadata": metadata_path.name,
                    }
                )
                png_temporary = png_path.with_suffix(
                    png_path.suffix + f".{uuid4().hex}.tmp"
                )
                try:
                    png_temporary.write_bytes(capture.png)
                    png_temporary.replace(png_path)
                    write_json(metadata, metadata_path)
                finally:
                    png_temporary.unlink(missing_ok=True)
            image_url = "data:image/png;base64," + base64.b64encode(
                capture.png
            ).decode("ascii")
            return [
                {
                    "type": "text",
                    "text": json.dumps(metadata, sort_keys=True),
                },
                {
                    "type": "image_url",
                    "image_url": {"url": image_url, "detail": "original"},
                },
            ]

        return StructuredTool.from_function(
            func=inspect_workbook_view,
            name="inspect_workbook_view",
            description=(
                "Open a supplied input or complete Excel workbook read-only, show "
                "an exact worksheet viewport through the configured LibreOffice "
                "backend, and return viewport metadata plus a PNG screenshot."
            ),
            handle_tool_error=True,
        )

    def close(self) -> None:
        if self._backend is not None:
            self._backend.close()
            self._backend = None


def create_visual_runtime(
    workbooks: dict[str, Path],
    *,
    allowed_sheets: set[str] | None = None,
    artifacts_dir: Path | None = None,
    config: VisualWorkbookConfig | None = None,
) -> VisualWorkbookRuntime:
    """Create a lazy stage-scoped runtime from explicit config or environment."""

    return VisualWorkbookRuntime(
        config or VisualWorkbookConfig.from_env(),
        workbooks,
        allowed_sheets=allowed_sheets,
        artifacts_dir=artifacts_dir,
    )
