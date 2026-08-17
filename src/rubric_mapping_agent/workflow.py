"""Standalone Part 1-3 functions, full pipeline, and minimal CLI."""

from __future__ import annotations

import argparse
import json
import logging
import os
import shutil
import tempfile
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4

from openai import OpenAI
from openpyxl import load_workbook
from rubric_mapping_eval.i2c_mapping import parse_item_mapping, parse_rubric
from rubric_mapping_eval.sectioning import parse_sections

from .agent import PROJECT_ROOT, StageResponse, build_agent


SKILL_SOURCE = PROJECT_ROOT / "skills" / "xlsx-rubric-mapping"
LOGGER = logging.getLogger(__name__)
CODE_INTERPRETER_MEMORY_LIMITS = {"1g", "4g", "16g", "64g"}


def _require_files(paths: Iterable[Path]) -> None:
    missing = [path for path in paths if not path.is_file()]
    if missing:
        raise FileNotFoundError("Missing input file(s): " + ", ".join(map(str, missing)))


def _stage_prompt(
    stage: str,
    local_files: dict[str, Path],
    python_files: dict[str, str],
) -> str:
    listed = "\n".join(
        f"- {name}: agent=/{local_files[name].as_posix()}; "
        f"python={python_files[name]}"
        for name in local_files
    )
    references = {
        "part1": (
            "workbook-inspection.md, part-1-overall-sections.md, "
            "and output-contracts.md"
        ),
        "part2": (
            "workbook-inspection.md, part-2-intermediate-sections.md, "
            "and output-contracts.md"
        ),
        "part3": (
            "workbook-inspection.md, part-3-items-to-cells.md, "
            "and output-contracts.md"
        ),
    }[stage]
    root_key = {"part1": "sections", "part2": "subsections", "part3": "items"}[stage]
    return f"""TASK_STAGE: {stage}

Use the xlsx-rubric-mapping skill. After reading SKILL.md, load only
{references} from that skill.

Allowed task files:
{listed}

Use the hosted python tool at least once. Its paths above are the authoritative
paths for Python code. Agent paths are only for read_file access when useful.
Do not inspect any path not listed above. Return an artifact whose only root key
is `{root_key}`. Do not write or modify files.
"""


def _memory_limit() -> str:
    memory_limit = os.getenv("OPENAI_CODE_INTERPRETER_MEMORY", "4g")
    if memory_limit not in CODE_INTERPRETER_MEMORY_LIMITS:
        choices = ", ".join(sorted(CODE_INTERPRETER_MEMORY_LIMITS))
        raise ValueError(f"OPENAI_CODE_INTERPRETER_MEMORY must be one of: {choices}")
    return memory_limit


def _invoke_stage(stage: str, sources: dict[str, Path]) -> dict[str, Any]:
    _require_files(sources.values())
    with tempfile.TemporaryDirectory(prefix=f"rubric-map-{stage}-") as temp_dir:
        workspace = Path(temp_dir)
        staged: dict[str, Path] = {}
        for name, source in sources.items():
            target = workspace / "task" / f"{name}{''.join(source.suffixes)}"
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
            staged[name] = target.relative_to(workspace)

        shutil.copytree(SKILL_SOURCE, workspace / "skills" / SKILL_SOURCE.name)
        client = OpenAI()
        container = client.containers.create(
            name=f"rubric-map-{stage}-{uuid4().hex[:8]}",
            memory_limit=_memory_limit(),
            network_policy={"type": "disabled"},
        )
        try:
            python_files: dict[str, str] = {}
            for name, relative_path in staged.items():
                with (workspace / relative_path).open("rb") as upload:
                    remote_file = client.containers.files.create(
                        container.id,
                        file=upload,
                    )
                python_files[name] = remote_file.path

            agent = build_agent(workspace, container.id)
            result = agent.invoke(
                {
                    "messages": [
                        {
                            "role": "user",
                            "content": _stage_prompt(stage, staged, python_files),
                        }
                    ]
                },
                config={"configurable": {"thread_id": f"{stage}-{uuid4()}"}},
            )
        finally:
            try:
                client.containers.delete(container.id)
            except Exception:
                LOGGER.warning(
                    "Could not delete expired or unreachable Code Interpreter container %s",
                    container.id,
                    exc_info=True,
                )

    response = result.get("structured_response")
    if isinstance(response, StageResponse):
        return response.artifact
    if isinstance(response, dict) and isinstance(response.get("artifact"), dict):
        return response["artifact"]
    raise ValueError(f"{stage} agent did not return a structured artifact")


def _write_json(payload: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_suffix(output_path.suffix + f".{uuid4().hex}.tmp")
    temporary.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    temporary.replace(output_path)


def _validate_subsections(
    payload: dict[str, Any], sections_path: Path | None = None
) -> None:
    if (
        set(payload) != {"subsections"}
        or not isinstance(payload["subsections"], list)
        or not payload["subsections"]
    ):
        raise ValueError("subsections.json must contain only a subsections list")
    parents = None
    if sections_path is not None:
        sections = parse_sections(json.loads(sections_path.read_text(encoding="utf-8")))
        parents = {section.section_id: section for section in sections}
    seen: set[str] = set()
    required = {"subsection_id", "parent_section_id", "sheet", "cells", "roles"}
    for index, subsection in enumerate(payload["subsections"]):
        if not isinstance(subsection, dict) or set(subsection) != required:
            raise ValueError(f"subsections[{index}] must contain exactly {sorted(required)}")
        identifier = subsection["subsection_id"]
        if not isinstance(identifier, str) or not identifier or identifier in seen:
            raise ValueError(f"invalid or duplicate subsection_id at index {index}")
        seen.add(identifier)
        for field in ("parent_section_id", "sheet"):
            if not isinstance(subsection[field], str) or not subsection[field]:
                raise ValueError(f"subsections[{index}].{field} must be non-empty")
        if (
            not isinstance(subsection["cells"], list)
            or not subsection["cells"]
            or not all(isinstance(value, str) and value for value in subsection["cells"])
        ):
            raise ValueError(f"subsections[{index}].cells must contain addresses")
        if (
            not isinstance(subsection["roles"], list)
            or not subsection["roles"]
            or not all(isinstance(value, str) and value for value in subsection["roles"])
        ):
            raise ValueError(f"subsections[{index}].roles must contain semantic tags")
        if parents is not None:
            parent = parents.get(subsection["parent_section_id"])
            if parent is None or parent.sheet != subsection["sheet"]:
                raise ValueError(f"subsections[{index}] has an invalid Part 1 parent")
            parent_addresses = {cell.address for cell in parent.cells}
            if not set(subsection["cells"]).issubset(parent_addresses):
                raise ValueError(f"subsections[{index}] contains cells outside its parent")


def _eligible_diff_cells(input_path: Path, complete_path: Path) -> set[tuple[str, str]]:
    initial = load_workbook(input_path, data_only=False, read_only=False)
    complete = load_workbook(complete_path, data_only=False, read_only=False)
    try:
        eligible: set[tuple[str, str]] = set()
        for completed_sheet in complete.worksheets:
            initial_sheet = (
                initial[completed_sheet.title]
                if completed_sheet.title in initial.sheetnames
                else None
            )
            for row in completed_sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    initial_value = (
                        initial_sheet[cell.coordinate].value
                        if initial_sheet is not None
                        else None
                    )
                    if initial_value != cell.value:
                        eligible.add((completed_sheet.title, cell.coordinate.upper()))
        return eligible
    finally:
        initial.close()
        complete.close()


def create_overall_section(
    input_path: str | Path,
    complete_path: str | Path,
    instructions_path: str | Path,
    *,
    output_path: str | Path = "sections.json",
) -> dict[str, Any]:
    """Create the assignment-required Part 1 artifact without rubric access."""

    sources = {
        "input": Path(input_path).resolve(),
        "complete": Path(complete_path).resolve(),
        "instructions": Path(instructions_path).resolve(),
    }
    artifact = _invoke_stage("part1", sources)
    parse_sections(artifact)
    _write_json(artifact, Path(output_path).resolve())
    return artifact


def create_intermediate_sections(
    input_path: str | Path,
    complete_path: str | Path,
    instructions_path: str | Path,
    sections_path: str | Path,
    *,
    output_path: str | Path = "subsections.json",
) -> dict[str, Any]:
    """Create the project-owned, rubric-free Part 2 handoff artifact."""

    sources = {
        "input": Path(input_path).resolve(),
        "complete": Path(complete_path).resolve(),
        "instructions": Path(instructions_path).resolve(),
        "sections": Path(sections_path).resolve(),
    }
    _require_files(sources.values())
    parse_sections(json.loads(sources["sections"].read_text(encoding="utf-8")))
    artifact = _invoke_stage("part2", sources)
    _validate_subsections(artifact, sources["sections"])
    _write_json(artifact, Path(output_path).resolve())
    return artifact


def create_items_to_cells_mapping(
    input_path: str | Path,
    complete_path: str | Path,
    instructions_path: str | Path,
    rubric_path: str | Path,
    *,
    sections_path: str | Path | None = None,
    subsections_path: str | Path | None = None,
    output_path: str | Path = "items_to_cells.json",
) -> dict[str, Any]:
    """Create the assignment-required Part 3 artifact."""

    input_file = Path(input_path).resolve()
    complete_file = Path(complete_path).resolve()
    rubric_file = Path(rubric_path).resolve()
    sources = {
        "input": input_file,
        "complete": complete_file,
        "instructions": Path(instructions_path).resolve(),
        "rubric": rubric_file,
    }
    if sections_path is not None:
        sources["sections"] = Path(sections_path).resolve()
    if subsections_path is not None:
        if sections_path is None:
            raise ValueError("subsections_path requires sections_path")
        sources["subsections"] = Path(subsections_path).resolve()

    _require_files(sources.values())
    criteria = parse_rubric(json.loads(rubric_file.read_text(encoding="utf-8")))
    if "sections" in sources:
        parse_sections(json.loads(sources["sections"].read_text(encoding="utf-8")))
    if "subsections" in sources:
        _validate_subsections(
            json.loads(sources["subsections"].read_text(encoding="utf-8")),
            sources["sections"],
        )

    artifact = _invoke_stage("part3", sources)
    parsed = parse_item_mapping(artifact, allow_empty_cells=True)
    expected_ids = {item for criterion in criteria for item in criterion.item_ids}
    if set(parsed) != expected_ids:
        raise ValueError("items_to_cells.json item IDs do not match rubric.json")

    eligible = _eligible_diff_cells(input_file, complete_file)
    invalid = sorted(
        (cell.sheet, cell.address)
        for cells in parsed.values()
        for cell in cells
        if (cell.sheet, cell.address) not in eligible
    )
    if invalid:
        raise ValueError(f"Part 3 mapped cells outside the eligible diff: {invalid[:10]}")

    _write_json(artifact, Path(output_path).resolve())
    return artifact


def run_complete_workflow(
    input_path: str | Path,
    complete_path: str | Path,
    instructions_path: str | Path,
    rubric_path: str | Path,
    *,
    output_dir: str | Path,
) -> dict[str, Path]:
    """Run Parts 1, 2, and 3 as separate agent invocations."""

    output = Path(output_dir).resolve()
    sections = output / "sections.json"
    subsections = output / "subsections.json"
    items = output / "items_to_cells.json"
    create_overall_section(
        input_path, complete_path, instructions_path, output_path=sections
    )
    create_intermediate_sections(
        input_path,
        complete_path,
        instructions_path,
        sections,
        output_path=subsections,
    )
    create_items_to_cells_mapping(
        input_path,
        complete_path,
        instructions_path,
        rubric_path,
        sections_path=sections,
        subsections_path=subsections,
        output_path=items,
    )
    return {"sections": sections, "subsections": subsections, "items_to_cells": items}


def _add_common(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--complete", required=True, type=Path)
    parser.add_argument("--instructions", required=True, type=Path)


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run the rubric-mapping agent workflow")
    subparsers = parser.add_subparsers(dest="command", required=True)

    part1 = subparsers.add_parser("part1", help="create sections.json")
    _add_common(part1)
    part1.add_argument("--output", type=Path, default=Path("sections.json"))

    part2 = subparsers.add_parser("part2", help="create subsections.json")
    _add_common(part2)
    part2.add_argument("--sections", required=True, type=Path)
    part2.add_argument("--output", type=Path, default=Path("subsections.json"))

    part3 = subparsers.add_parser("part3", help="create items_to_cells.json")
    _add_common(part3)
    part3.add_argument("--rubric", required=True, type=Path)
    part3.add_argument("--sections", type=Path)
    part3.add_argument("--subsections", type=Path)
    part3.add_argument("--output", type=Path, default=Path("items_to_cells.json"))

    all_stages = subparsers.add_parser("all", help="run Parts 1, 2, and 3")
    _add_common(all_stages)
    all_stages.add_argument("--rubric", required=True, type=Path)
    all_stages.add_argument("--output-dir", required=True, type=Path)
    return parser


def main() -> int:
    args = _parser().parse_args()
    common = (args.input, args.complete, args.instructions)
    if args.command == "part1":
        create_overall_section(*common, output_path=args.output)
        print(args.output.resolve())
    elif args.command == "part2":
        create_intermediate_sections(
            *common, args.sections, output_path=args.output
        )
        print(args.output.resolve())
    elif args.command == "part3":
        create_items_to_cells_mapping(
            *common,
            args.rubric,
            sections_path=args.sections,
            subsections_path=args.subsections,
            output_path=args.output,
        )
        print(args.output.resolve())
    else:
        outputs = run_complete_workflow(
            *common, args.rubric, output_dir=args.output_dir
        )
        print(json.dumps({key: str(path) for key, path in outputs.items()}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
