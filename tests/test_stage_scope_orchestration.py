from __future__ import annotations

import json
import os
import tempfile
import threading
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from rubric_mapping_agent.workflow import (
    _combine_part2_artifacts,
    _combine_part3_artifacts,
    _invoke_sheets_in_parallel,
    create_intermediate_sections,
    create_items_to_cells_mapping,
    create_overall_section,
)


class StageScopeOrchestrationTests(unittest.TestCase):
    def _workbook(self, path: Path, sheets: tuple[str, ...]) -> None:
        workbook = Workbook()
        workbook.active.title = sheets[0]
        for sheet_name in sheets[1:]:
            workbook.create_sheet(sheet_name)
        workbook.save(path)
        workbook.close()

    def test_sheet_invocations_overlap_but_results_keep_workbook_order(self) -> None:
        barrier = threading.Barrier(2)

        def invoke(stage, sources, *, target_sheet, visual_artifacts_dir=None):
            self.assertEqual(stage, "part1")
            barrier.wait(timeout=5)
            return {"sheet": target_sheet}

        with patch.dict(
            os.environ,
            {"RUBRIC_MAP_SHEET_MAX_WORKERS": "2"},
            clear=True,
        ), patch(
            "rubric_mapping_agent.workflow._invoke_stage",
            side_effect=invoke,
        ):
            results = _invoke_sheets_in_parallel(
                "part1",
                ("Alpha", "Beta"),
                {"input": Path("input.xlsx")},
            )

        self.assertEqual(
            results,
            (
                ("Alpha", {"sheet": "Alpha"}),
                ("Beta", {"sheet": "Beta"}),
            ),
        )

    def test_part1_workbook_scope_uses_one_invocation(self) -> None:
        artifact = {
            "sections": [
                {"section_id": "local_a", "sheet": "Alpha", "cells": ["A1"]},
                {"section_id": "local_b", "sheet": "Beta", "cells": ["B2"]},
            ],
            "section_summaries": [
                {
                    "section_id": "local_a",
                    "title": "Alpha Panel",
                    "detail": "Contains the Alpha worksheet panel.",
                    "plain_language": "Explains the Alpha model area.",
                },
                {
                    "section_id": "local_b",
                    "title": "Beta Panel",
                    "detail": "Contains the Beta worksheet panel.",
                    "plain_language": "Explains the Beta model area.",
                },
            ],
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "input.xlsx"
            complete_path = root / "complete.xlsx"
            self._workbook(input_path, ("Alpha", "Beta"))
            self._workbook(complete_path, ("Alpha", "Beta"))
            instructions = root / "instructions.md"
            instructions.write_text("Build the model.", encoding="utf-8")
            output = root / "sections.json"
            with patch.dict(
                os.environ,
                {"RUBRIC_MAP_PART1_SCOPE": "workbook"},
                clear=True,
            ), patch(
                "rubric_mapping_agent.workflow._invoke_stage",
                return_value=artifact,
            ) as invoke:
                result = create_overall_section(
                    input_path,
                    complete_path,
                    instructions,
                    output_path=output,
                )

        invoke.assert_called_once()
        self.assertNotIn("target_sheet", invoke.call_args.kwargs)
        self.assertEqual(
            [section["section_id"] for section in result["sections"]],
            ["section_001", "section_002"],
        )

    def test_part2_sheet_scope_concatenates_unique_ids_in_workbook_order(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "input.xlsx"
            complete_path = root / "complete.xlsx"
            self._workbook(input_path, ("Alpha", "Beta"))
            self._workbook(complete_path, ("Alpha", "Beta"))
            completed = load_workbook(complete_path)
            completed["Alpha"]["A1"] = 1
            completed["Beta"]["B2"] = 2
            completed.save(complete_path)
            completed.close()
            instructions = root / "instructions.md"
            instructions.write_text("Build the model.", encoding="utf-8")
            sections = root / "sections.json"
            sections.write_text(
                json.dumps(
                    {
                        "sections": [
                            {
                                "section_id": "section_001",
                                "sheet": "Alpha",
                                "cells": ["A1"],
                            },
                            {
                                "section_id": "section_002",
                                "sheet": "Beta",
                                "cells": ["B2"],
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )
            section_summary = root / "part1-summary.md"
            section_summary.write_text(
                "# Part 1 Section Summary\n\n"
                "## section_001 — Alpha Panel\n\n"
                "**Detail:** Contains the Alpha model panel.\n\n"
                "**In normal words:** Shows the Alpha model area.\n\n"
                "**Worksheet:** `Alpha`\n\n"
                "## section_002 — Beta Panel\n\n"
                "**Detail:** Contains the Beta model panel.\n\n"
                "**In normal words:** Shows the Beta model area.\n\n"
                "**Worksheet:** `Beta`\n",
                encoding="utf-8",
            )
            output = root / "subsections.json"
            calls: list[str] = []

            def invoke(stage, sources, *, target_sheet, visual_artifacts_dir=None):
                self.assertEqual(stage, "part2")
                calls.append(target_sheet)
                parent = "section_001" if target_sheet == "Alpha" else "section_002"
                address = "A1" if target_sheet == "Alpha" else "B2"
                identifier = (
                    "subsection_s001_001"
                    if target_sheet == "Alpha"
                    else "subsection_s002_001"
                )
                return {
                    "subsections": [
                        {
                            "subsection_id": identifier,
                            "parent_section_id": parent,
                            "sheet": target_sheet,
                            "cells": [address],
                            "roles": ["output"],
                        }
                    ],
                    "subsection_index": {
                        "schema_version": 2,
                        "generated_by": "part2_agent",
                        "families": [
                            {
                                "family_id": f"{identifier}_family_01",
                                "subsection_id": identifier,
                                "parent_section_id": parent,
                                "sheet": target_sheet,
                                "object_name": f"{target_sheet} Output",
                                "aliases": [],
                                "changed_cells": [address],
                                "anchor_cells": [],
                                "roles": ["output"],
                                "scope": {
                                    "period_type": "unspecified",
                                    "period_headers": [],
                                },
                                "orientation": "row",
                                "calculation_kind": "output",
                                "formula_signatures": [],
                            }
                        ],
                        "relationships": [],
                    },
                }

            with patch.dict(
                os.environ,
                {
                    "RUBRIC_MAP_PART2_SCOPE": "sheet",
                    "RUBRIC_MAP_SHEET_MAX_WORKERS": "2",
                },
                clear=True,
            ), patch(
                "rubric_mapping_agent.workflow._invoke_stage",
                side_effect=invoke,
            ):
                result = create_intermediate_sections(
                    input_path,
                    complete_path,
                    instructions,
                    sections,
                    section_summary,
                    output_path=output,
                )
                index_payload = json.loads(
                    output.with_name("subsection_index.json").read_text(
                        encoding="utf-8"
                    )
                )

        self.assertCountEqual(calls, ["Alpha", "Beta"])
        self.assertEqual(
            [item["subsection_id"] for item in result["subsections"]],
            ["subsection_s001_001", "subsection_s002_001"],
        )
        self.assertEqual(
            [item["sheet"] for item in result["subsections"]],
            ["Alpha", "Beta"],
        )
        self.assertEqual(
            [family["family_id"] for family in index_payload["families"]],
            [
                "subsection_s001_001_family_01",
                "subsection_s002_001_family_01",
            ],
        )

    def test_part3_sheet_scope_unions_cells_for_each_rubric_item(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "input.xlsx"
            complete_path = root / "complete.xlsx"
            self._workbook(input_path, ("Alpha", "Beta"))
            self._workbook(complete_path, ("Alpha", "Beta"))
            complete = Workbook()
            complete.active.title = "Alpha"
            complete["Alpha"]["A1"] = 1
            complete.create_sheet("Beta")["B2"] = 2
            complete.save(complete_path)
            complete.close()
            instructions = root / "instructions.md"
            instructions.write_text("Build the model.", encoding="utf-8")
            rubric = root / "rubric.json"
            rubric.write_text(
                json.dumps(
                    {
                        "criteria": {
                            "criterion_1": {
                                "criterion_id": 1,
                                "grading": [{"item_id": "1.1"}],
                            }
                        }
                    }
                ),
                encoding="utf-8",
            )
            output = root / "items_to_cells.json"
            calls: list[str] = []

            def invoke(stage, sources, *, target_sheet, visual_artifacts_dir=None):
                self.assertEqual(stage, "part3")
                calls.append(target_sheet)
                address = "A1" if target_sheet == "Alpha" else "B2"
                return {
                    "items": [
                        {
                            "item_id": "1.1",
                            "cells": [
                                {"sheet": target_sheet, "address": address}
                            ],
                        }
                    ]
                }

            with patch.dict(
                os.environ,
                {
                    "RUBRIC_MAP_PART3_SCOPE": "sheet",
                    "RUBRIC_MAP_SHEET_MAX_WORKERS": "2",
                },
                clear=True,
            ), patch(
                "rubric_mapping_agent.workflow._invoke_stage",
                side_effect=invoke,
            ):
                result = create_items_to_cells_mapping(
                    input_path,
                    complete_path,
                    instructions,
                    rubric,
                    output_path=output,
                )

        self.assertCountEqual(calls, ["Alpha", "Beta"])
        self.assertEqual(
            result,
            {
                "items": [
                    {
                        "item_id": "1.1",
                        "cells": [
                            {"sheet": "Alpha", "address": "A1"},
                            {"sheet": "Beta", "address": "B2"},
                        ],
                    }
                ]
            },
        )

    def test_part2_sheet_scope_rejects_cross_sheet_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            sections = Path(temp_dir) / "sections.json"
            sections.write_text(
                json.dumps(
                    {
                        "sections": [
                            {
                                "section_id": "section_001",
                                "sheet": "Beta",
                                "cells": ["A1"],
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            artifact = {
                "subsections": [
                    {
                        "subsection_id": "subsection_s002_001",
                        "parent_section_id": "section_001",
                        "sheet": "Beta",
                        "cells": ["A1"],
                        "roles": ["output"],
                    }
                ],
                "subsection_index": {
                    "schema_version": 2,
                    "generated_by": "part2_agent",
                    "families": [],
                    "relationships": [],
                },
            }

            with self.assertRaisesRegex(ValueError, "another sheet"):
                _combine_part2_artifacts((("Alpha", artifact),), sections, set())

    def test_part2_sheet_scope_tolerates_missing_index_families(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            sections = Path(temp_dir) / "sections.json"
            sections.write_text(
                json.dumps(
                    {
                        "sections": [
                            {
                                "section_id": "section_001",
                                "sheet": "Alpha",
                                "cells": ["A1"],
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            artifact = {
                "subsections": [
                    {
                        "subsection_id": "subsection_s001_001",
                        "parent_section_id": "section_001",
                        "sheet": "Alpha",
                        "cells": ["A1"],
                        "roles": ["output"],
                    }
                ],
                "subsection_index": {
                    "schema_version": 2,
                    "generated_by": "part2_agent",
                    "relationships": [],
                },
            }

            subsections, index_payload = _combine_part2_artifacts(
                (("Alpha", artifact),), sections, {("Alpha", "A1")}
            )

        self.assertEqual(len(subsections["subsections"]), 1)
        self.assertEqual(index_payload["families"], [])
        self.assertEqual(index_payload["relationships"], [])

    def test_part3_sheet_scope_rejects_cross_sheet_output(self) -> None:
        artifact = {
            "items": [
                {
                    "item_id": "1.1",
                    "cells": [{"sheet": "Beta", "address": "A1"}],
                }
            ]
        }

        with self.assertRaisesRegex(ValueError, "another sheet"):
            _combine_part3_artifacts(
                (("Alpha", artifact),),
                expected_item_ids=("1.1",),
                eligible={("Beta", "A1")},
                workbook_sheet_order=("Alpha", "Beta"),
            )


if __name__ == "__main__":
    unittest.main()
