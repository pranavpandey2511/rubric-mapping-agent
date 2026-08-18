from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from rubric_mapping_agent.review import (
    COMMENT_MARKER,
    LEGEND_MARKER,
    PART1_OUTLINE_COLOR,
    PART2_HISTORICAL_COLOR,
    PART2_PROJECTED_COLOR,
    PART3_HIGHLIGHT_COLOR,
    create_review_workbook,
    visualize_mapping_outputs,
)


class ReviewWorkbookTests(unittest.TestCase):
    def _write_json(self, path: Path, payload: dict) -> None:
        path.write_text(json.dumps(payload), encoding="utf-8")

    def _fixture(self, root: Path) -> tuple[Path, Path, Path, Path, Path]:
        workbook_path = root / "model.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Model"
        for row in range(1, 5):
            for column in range(1, 5):
                sheet.cell(row=row, column=column, value=row * column)
        sheet["B2"].font = Font(bold=True, color="112233")
        sheet["C3"].fill = PatternFill(fill_type="solid", fgColor="ABCDEF")
        sheet["C3"].comment = Comment("Existing reviewer note", "Reviewer")
        workbook.save(workbook_path)

        sections_path = root / "sections.json"
        self._write_json(
            sections_path,
            {
                "sections": [
                    {
                        "section_id": "section_001",
                        "sheet": "Model",
                        "cells": ["B2", "C2", "B3", "C3"],
                    }
                ]
            },
        )
        subsections_path = root / "subsections.json"
        self._write_json(
            subsections_path,
            {
                "subsections": [
                    {
                        "subsection_id": "subsection_001",
                        "parent_section_id": "section_001",
                        "sheet": "Model",
                        "cells": ["B2", "C2"],
                        "roles": ["historical", "input"],
                    }
                ]
            },
        )
        items_path = root / "items_to_cells.json"
        self._write_json(
            items_path,
            {
                "items": [
                    {"item_id": "1.1", "cells": [{"sheet": "Model", "address": "C3"}]},
                    {"item_id": "1.2", "cells": [{"sheet": "Model", "address": "C3"}]},
                ]
            },
        )
        rubric_path = root / "rubric.json"
        self._write_json(
            rubric_path,
            {
                "criteria": {
                    "criterion_1": {
                        "criterion_id": 1,
                        "description": "Revenue build",
                        "grading": [
                            {"item_id": "1.1", "condition": "Link the historical value."},
                            {"item_id": "1.2", "condition": "Calculate the projected value."},
                        ],
                    }
                }
            },
        )
        return workbook_path, sections_path, subsections_path, items_path, rubric_path

    def test_visualizes_all_layers_without_overwriting_source(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path, sections, subsections, items, rubric = self._fixture(root)
            output = root / "annotated.xlsx"

            result = create_review_workbook(
                workbook_path,
                output,
                sections_path=sections,
                subsections_path=subsections,
                items_to_cells_path=items,
                rubric_path=rubric,
            )

            self.assertEqual(result, output.resolve())
            original = load_workbook(workbook_path)
            annotated = load_workbook(output)
            try:
                self.assertIsNone(original["Model"]["B2"].border.top.style)
                self.assertEqual(original["Model"]["C3"].comment.text, "Existing reviewer note")

                model = annotated["Model"]
                self.assertEqual(model["B2"].border.top.style, "medium")
                self.assertEqual(model["B2"].border.left.style, "medium")
                self.assertEqual(
                    model["B2"].border.top.color.rgb[-6:], PART1_OUTLINE_COLOR
                )
                self.assertEqual(model["C2"].border.bottom.style, "medium")
                self.assertEqual(
                    model["C2"].border.bottom.color.rgb[-6:],
                    PART2_HISTORICAL_COLOR,
                )
                self.assertEqual(model["B2"].font.color.rgb[-6:], "112233")
                self.assertEqual(model["C3"].fill.fill_type, "solid")
                self.assertEqual(
                    model["C3"].fill.fgColor.rgb[-6:], PART3_HIGHLIGHT_COLOR
                )
                self.assertIn("Existing reviewer note", model["C3"].comment.text)
                self.assertIn(COMMENT_MARKER, model["C3"].comment.text)
                self.assertIn("1.1", model["C3"].comment.text)
                self.assertIn("1.2", model["C3"].comment.text)
                self.assertEqual(annotated["Mapping Legend"]["A1"].value, LEGEND_MARKER)
            finally:
                original.close()
                annotated.close()

    def test_rejects_subsection_cell_outside_parent(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path, sections, subsections, _, _ = self._fixture(root)
            payload = json.loads(subsections.read_text(encoding="utf-8"))
            payload["subsections"][0]["cells"] = ["A1"]
            self._write_json(subsections, payload)

            with self.assertRaisesRegex(ValueError, "outside its parent"):
                visualize_mapping_outputs(
                    workbook_path,
                    root / "annotated.xlsx",
                    sections_path=sections,
                    subsections_path=subsections,
                )

    def test_uses_uniform_section_outlines_and_distinct_subsection_borders(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path, sections, subsections, _, _ = self._fixture(root)
            section_payload = json.loads(sections.read_text(encoding="utf-8"))
            section_payload["sections"][0]["cells"].extend(["B4", "C4"])
            section_payload["sections"].append(
                {
                    "section_id": "section_002",
                    "sheet": "Model",
                    "cells": ["A1", "A2"],
                }
            )
            self._write_json(sections, section_payload)
            subsection_payload = json.loads(subsections.read_text(encoding="utf-8"))
            subsection_payload["subsections"].append(
                {
                    "subsection_id": "subsection_002",
                    "parent_section_id": "section_001",
                    "sheet": "Model",
                    "cells": ["B3", "C3"],
                    "roles": ["projected", "output"],
                }
            )
            self._write_json(subsections, subsection_payload)

            output = root / "annotated.xlsx"
            visualize_mapping_outputs(
                workbook_path,
                output,
                sections_path=sections,
                subsections_path=subsections,
                include_legend=False,
            )

            annotated = load_workbook(output)
            try:
                model = annotated["Model"]
                self.assertEqual(
                    model["A1"].border.top.color.rgb[-6:], PART1_OUTLINE_COLOR
                )
                self.assertEqual(
                    model["B2"].border.top.color.rgb[-6:], PART1_OUTLINE_COLOR
                )
                self.assertEqual(model["B2"].border.bottom.style, "medium")
                self.assertEqual(model["B3"].border.bottom.style, "medium")
                self.assertEqual(
                    model["B2"].border.bottom.color.rgb[-6:],
                    PART2_HISTORICAL_COLOR,
                )
                self.assertEqual(
                    model["B3"].border.bottom.color.rgb[-6:],
                    PART2_PROJECTED_COLOR,
                )
            finally:
                annotated.close()

    def test_draws_each_time_series_row_with_historical_and_projected_boundaries(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "model.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Model"
            sheet["C1"] = "Revenue Projection"
            sheet.merge_cells("D1:F1")
            sheet["D1"] = "Historical"
            sheet.merge_cells("G1:H1")
            sheet["G1"] = "Projected"
            sheet.append([None, None, None, "FY21", "FY22", "FY23", "FY24", "FY25"])
            sheet.append([None, None, "Revenue", 10, 20, 30, 40, 50])
            sheet.append([None, None, "% growth", None, 0.1, 0.2, 0.3, 0.4])
            workbook.save(workbook_path)
            workbook.close()

            sections = root / "sections.json"
            self._write_json(
                sections,
                {
                    "sections": [
                        {
                            "section_id": "section_001",
                            "sheet": "Model",
                            "cells": [
                                f"{column}{row}"
                                for row in range(1, 6)
                                for column in "CDEFGH"
                            ],
                        }
                    ]
                },
            )
            subsections = root / "subsections.json"
            self._write_json(
                subsections,
                {
                    "subsections": [
                        {
                            "subsection_id": "subsection_001",
                            "parent_section_id": "section_001",
                            "sheet": "Model",
                            "cells": [
                                f"{column}{row}"
                                for row in range(1, 5)
                                for column in "CDEFGH"
                            ],
                            "roles": ["historical-and-projected"],
                        }
                    ]
                },
            )

            output = root / "annotated.xlsx"
            visualize_mapping_outputs(
                workbook_path,
                output,
                sections_path=sections,
                subsections_path=subsections,
                include_legend=False,
            )

            annotated = load_workbook(output)
            try:
                model = annotated["Model"]
                for address in ("D2", "E2", "F2", "D3", "E3", "F3"):
                    self.assertEqual(model[address].border.bottom.style, "medium")
                    self.assertEqual(
                        model[address].border.bottom.color.rgb[-6:],
                        PART2_HISTORICAL_COLOR,
                    )
                for address in ("G2", "H2", "G3", "H3", "G4", "H4"):
                    self.assertEqual(model[address].border.bottom.style, "medium")
                    self.assertEqual(
                        model[address].border.bottom.color.rgb[-6:],
                        PART2_PROJECTED_COLOR,
                    )
                self.assertIsNone(model["D4"].border.bottom.style)
                self.assertEqual(
                    model["E4"].border.bottom.color.rgb[-6:],
                    PART2_HISTORICAL_COLOR,
                )
                self.assertIsNone(model["C3"].border.bottom.style)
                self.assertEqual(model["C5"].border.bottom.style, "medium")
                self.assertEqual(
                    model["C5"].border.bottom.color.rgb[-6:], PART1_OUTLINE_COLOR
                )
            finally:
                annotated.close()

    def test_extends_part2_visualization_to_uncovered_sections_and_sheets(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path, sections, subsections, _, _ = self._fixture(root)
            workbook = load_workbook(workbook_path)
            second = workbook.create_sheet("Second")
            second["C1"] = "Revenue Projection"
            second.merge_cells("D1:F1")
            second["D1"] = "Historical"
            second.merge_cells("G1:H1")
            second["G1"] = "Projected"
            second.append([None, None, None, "FY21", "FY22", "FY23", "FY24", "FY25"])
            second.append([None, None, "Revenue", 10, 20, 30, 40, 50])
            workbook.save(workbook_path)
            workbook.close()

            section_payload = json.loads(sections.read_text(encoding="utf-8"))
            section_payload["sections"].append(
                {
                    "section_id": "section_002",
                    "sheet": "Second",
                    "cells": [
                        f"{column}{row}"
                        for row in range(1, 5)
                        for column in "CDEFGH"
                    ],
                }
            )
            self._write_json(sections, section_payload)

            output = root / "annotated.xlsx"
            visualize_mapping_outputs(
                workbook_path,
                output,
                sections_path=sections,
                subsections_path=subsections,
            )

            annotated = load_workbook(output)
            try:
                second = annotated["Second"]
                self.assertEqual(
                    second["D2"].border.bottom.color.rgb[-6:],
                    PART2_HISTORICAL_COLOR,
                )
                self.assertEqual(
                    second["G2"].border.bottom.color.rgb[-6:],
                    PART2_PROJECTED_COLOR,
                )
                self.assertEqual(
                    second["D3"].border.bottom.color.rgb[-6:],
                    PART2_HISTORICAL_COLOR,
                )
                self.assertEqual(
                    second["G3"].border.bottom.color.rgb[-6:],
                    PART2_PROJECTED_COLOR,
                )
                self.assertIsNone(second["C3"].border.bottom.style)
                legend_ids = {
                    cell.value
                    for cell in annotated["Mapping Legend"]["C"]
                    if cell.value is not None
                }
                self.assertIn("visual_section_002", legend_ids)
            finally:
                annotated.close()

    def test_requires_a_separate_output_path(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path, sections, _, _, _ = self._fixture(root)
            with self.assertRaisesRegex(ValueError, "must differ"):
                visualize_mapping_outputs(
                    workbook_path,
                    workbook_path,
                    sections_path=sections,
                )

    def test_consolidates_comments_at_a_merged_cell_anchor(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path, _, _, items, _ = self._fixture(root)
            workbook = load_workbook(workbook_path)
            workbook["Model"].merge_cells("B2:C2")
            workbook.save(workbook_path)
            workbook.close()
            self._write_json(
                items,
                {
                    "items": [
                        {
                            "item_id": "1.1",
                            "cells": [{"sheet": "Model", "address": "B2"}],
                        },
                        {
                            "item_id": "1.2",
                            "cells": [{"sheet": "Model", "address": "C2"}],
                        },
                    ]
                },
            )

            output = root / "annotated.xlsx"
            visualize_mapping_outputs(
                workbook_path,
                output,
                items_to_cells_path=items,
                include_legend=False,
            )

            annotated = load_workbook(output)
            try:
                comment = annotated["Model"]["B2"].comment
                self.assertIn("1.1", comment.text)
                self.assertIn("1.2", comment.text)
                self.assertEqual(
                    annotated["Model"]["B2"].fill.fgColor.rgb[-6:],
                    PART3_HIGHLIGHT_COLOR,
                )
            finally:
                annotated.close()

    def test_preserves_cached_formula_results(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path, sections, _, _, _ = self._fixture(root)
            workbook = load_workbook(workbook_path)
            workbook["Model"]["A1"] = "=1+1"
            workbook["Model"]["D4"] = '="cached text"'
            workbook.save(workbook_path)
            workbook.close()

            patched = root / "cached.xlsx"
            with ZipFile(workbook_path, "r") as source, ZipFile(
                patched, "w", compression=ZIP_DEFLATED
            ) as destination:
                for member in source.infolist():
                    data = source.read(member)
                    if member.filename == "xl/worksheets/sheet1.xml":
                        data = data.replace(b"<v />", b"<v>2</v>", 1)
                        data = data.replace(
                            b'<c r="D4"><f>"cached text"</f><v /></c>',
                            b'<c r="D4" t="str"><f>"cached text"</f>'
                            b"<v>cached text</v></c>",
                        )
                        data = data.replace(
                            b'<c r="D3" t="n"><v>12</v></c>',
                            b'<c r="D3" t="n"><v>201.02081816859607</v></c>',
                        )
                    destination.writestr(member, data)
            patched.replace(workbook_path)

            output = root / "annotated.xlsx"
            visualize_mapping_outputs(
                workbook_path,
                output,
                sections_path=sections,
                include_legend=False,
            )

            formula_view = load_workbook(output, data_only=False)
            cached_view = load_workbook(output, data_only=True)
            try:
                self.assertEqual(formula_view["Model"]["A1"].value, "=1+1")
                self.assertEqual(cached_view["Model"]["A1"].value, 2)
                self.assertEqual(
                    formula_view["Model"]["D4"].value, '="cached text"'
                )
                self.assertEqual(cached_view["Model"]["D4"].value, "cached text")
                self.assertEqual(
                    cached_view["Model"]["D3"].value,
                    201.02081816859607,
                )
            finally:
                formula_view.close()
                cached_view.close()

            with ZipFile(output) as archive:
                worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
            self.assertIn(b"<v>201.02081816859607</v>", worksheet_xml)


if __name__ == "__main__":
    unittest.main()
